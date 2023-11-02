import openpyxl
import pomocne_funkce

# seznam názvů zpracovávaných souborů (takto musí být ve stejné složce jako náš .py soubor)
vsechny_soubory = ["Spol1.xlsx", "Spol2.xlsx", "Spol3.xlsx", "Spol4.xlsx", "Spol5.xlsx"]

pocet_radku = 5
pocet_sloupecku = 4

vysledky = pomocne_funkce.tabulka_vysledku(radku=pocet_radku, sloupcu=pocet_sloupecku)

# soucet bunek
for soubor_nazev in vsechny_soubory:
    excel_soubor = openpyxl.load_workbook(soubor_nazev, data_only=True)  # otevření/načtení souboru
    aktivni_list = excel_soubor['Hmotny_majetek']  # výběr aktivního listu



    for radek in range(pocet_radku):
        for sloupec in range(pocet_sloupecku):
            bunka = aktivni_list.cell(row=radek+2, column=sloupec+2)  # výběr buňky
            print(f"{radek+2}:{sloupec+2}: {bunka.value}")
            vysledky[radek][sloupec] += bunka.value # scitam s předchozími hodnotami

# načtení souboru/listu/buňky do kterého zapíšeme výsledek
nazev_vysledek = "Vysledek.xlsx"
excel_vysledek = openpyxl.load_workbook(nazev_vysledek) # otevření/načtení souboru - soubor již musí existovat
list_vysledek = excel_vysledek.active # výběr aktivního listu

for radek in range(pocet_radku):
    for sloupec in range(pocet_sloupecku):
        bunka_vysledek = list_vysledek.cell(row=radek+2, column=sloupec+2) # pracujeme s buňkou B 2
        bunka_vysledek.value = vysledky[radek][sloupec] # zápis nasčítané hodnoty do buňky ve výsledném souboru


excel_soubor = openpyxl.load_workbook(vsechny_soubory[0], data_only=True)  # otevření/načtení souboru
aktivni_list = excel_soubor.active
for sloupec in range(1, pocet_sloupecku+2):
    list_vysledek.cell(row=1, column=sloupec).value =\
        aktivni_list.cell(row=1, column=sloupec).value


excel_vysledek.save(nazev_vysledek) # uložení změn v excelu
