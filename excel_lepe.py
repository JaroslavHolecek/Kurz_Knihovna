import numpy as np
import openpyxl

def najdi_bunku_s_barevnym_pozadim(aktivni_list, prvni):
    # Procházení buněk na listu
    if prvni:
        for radek in aktivni_list.iter_rows(max_col=20, max_row=20):
            for bunka in radek:
                # Kontrola, zda má buňka barevné pozadí
                if bunka.fill.start_color.index != '00000000':
                    # Vracíme souřadnice buňky s barevným pozadím
                    return openpyxl.utils.coordinate_to_tuple(bunka.coordinate)
    else:
        for radek in reversed(list(aktivni_list.iter_rows(max_col=20, max_row=20))):
            for bunka in reversed(radek):
                if bunka.fill.start_color.index != '00000000':
                    return openpyxl.utils.coordinate_to_tuple(bunka.coordinate)

def nacti_data_z_excelu(nazev_souboru, nazev_listu):
    excel_soubor = openpyxl.load_workbook(nazev_souboru, data_only=True)  # otevření/načtení souboru
    aktivni_list = excel_soubor[nazev_listu]
    radek_start, sloupec_start = najdi_bunku_s_barevnym_pozadim(aktivni_list, prvni=True)
    radek_konec, sloupec_konec = najdi_bunku_s_barevnym_pozadim(aktivni_list, prvni=False)

    pocet_sloupcu = sloupec_konec - sloupec_start + 1
    pocet_radku = radek_konec - radek_start + 1

    jedna_vrstva = np.ones((pocet_radku, pocet_sloupcu), dtype=int)
    for index, hodnota in enumerate(
            aktivni_list.iter_rows(min_row=radek_start, max_row=radek_konec,
                                   min_col=sloupec_start, max_col=sloupec_konec,
                                   values_only=True)):
        jedna_vrstva[index] = np.array(hodnota)
    return jedna_vrstva


def uloz_2d_pole_do_excelu(pole, soubor, list, zacatecni_bunka="A1"):
    sesit = openpyxl.Workbook()
    list_vysledek = sesit.create_sheet(title=list)

    # Zjistíme pozici buňky (řádek a sloupec) na základě zadaného názvu
    radek, sloupec = openpyxl.utils.coordinate_to_tuple(zacatecni_bunka)

    # Vložení obsahu pole do excelovského listu od zadané buňky
    for i, radek_pole in enumerate(pole):
        for j, hodnota in enumerate(radek_pole):
            bunka = list_vysledek.cell(row=radek + i, column=sloupec + j, value=hodnota)

    # Uložení do souboru
    sesit.save(soubor)

vsechny_soubory = ("Spol1.xlsx", "Spol2.xlsx", "Spol3.xlsx", "Spol4.xlsx", "Spol5.xlsx")

data = nacti_data_z_excelu(vsechny_soubory[0], 'Hmotny_majetek')
for index, soubor in enumerate(vsechny_soubory[1:], start=1):
    data = np.dstack((data, nacti_data_z_excelu(soubor, 'Hmotny_majetek')))

soucet = data.sum(axis=0)
uloz_2d_pole_do_excelu(soucet, "VyslednyExcel.xlsx", "Hmotny_majetek", "C4")





